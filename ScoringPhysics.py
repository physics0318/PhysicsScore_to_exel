from openpyxl import Workbook, load_workbook
from tkinter import *
import tkinter.messagebox as msgbox

#GUI
root = Tk()
root.title("채점창")
root.geometry("233x130+100+100")
root.resizable(False, False)
root.wm_attributes("-topmost", 1)

#이름칸
lb1 = Label(root, text="이름")
txt1 = Entry(root)

#학번칸
lb2 = Label(root, text="학번")
txt2 = Entry(root)

#점수칸
lb3 = Label(root, text="점수")
txt3 = Entry(root)

#엑셀 없으면 만들고 있으면 불러오기
try:
    wb = load_workbook("실험수업점수표.xlsx")
    ws = wb.active
    if str(ws.cell(row=2,column=1).value) != "None":
        txt1.insert(0, str(ws.cell(row=2, column=1).value))
    if str(ws.cell(row=2,column=2).value) != "None":
        txt2.insert(0, str(ws.cell(row=2, column=2).value))
    if str(ws.cell(row=2,column=3).value) != "None":
        txt3.insert(0, str(ws.cell(row=2, column=3).value))

except:
    wb = Workbook()
    ws = wb.active
    ws.title = "실험수업점수표"

    ws.append(("이름", "학번", "1주차", "2주차", "3주차", "4주차", "5주차", "6주차", "7주차", "8주차", "9주차", "출석", "총점", "비고"))

#입력 학번에 해당하는 학생이 있는 행 찾기
def StuRow():
    stunum = txt2.get()
    count = 2
    for row in ws.iter_rows(min_row=2):
        if row[1].value == stunum:
            break
        else:
            count+=1
            pass
    if count == ws.max_row+1:
        return 1
    else:
        return count

#조회버튼 누르면
def find_score():
    if StuRow() == 1:
        msgbox.showwarning("경고","조회하신 학번의 학생은 학생 명단에 저장되지 않은 학생입니다.\n우선 점수를 먼저 저장해주세요.")
    else:
        txt3.delete(0, END)
        txt3.insert(0, str(ws.cell(row=StuRow(), column=w+2).value))
        txt1.delete(0, END)
        txt1.insert(0, str(ws.cell(row=StuRow(), column=1).value))
        lb_student.config(text=str(StuRow()-1)+"/"+str(ws.max_row-1)+" 번째 학생")

#입력버튼 누르면
def save_score():
    name = txt1.get()
    stunum = txt2.get()
    if StuRow() != 1:
        ws.cell(row=StuRow(), column=w+2, value=txt3.get())
    else:
        ws.append((name,stunum))
        ws.cell(row = ws.max_row, column=w+2, value=txt3.get())
    lb_student.config(text=str(StuRow()-1)+"/"+str(ws.max_row-1)+" 번째 학생")
    wb.save("실험수업점수표.xlsx")

#레포트 주차
w = 1
#레포트 번호 바꾸기 버튼 누르면
def weekp():
    global w
    w += 1
    if w == 10:
        w = 1
    lb_week.config(text=str(w)+"/9 주차 레포트")
    if StuRow() != 1:
        find_score()

def weekm():
    global w
    w -= 1
    if w == 0:
        w = 9
    lb_week.config(text=str(w)+"/9 주차 레포트")
    if StuRow() != 1:
        find_score()

#학생 바꾸기 버튼 누르면
def stup():
    s = StuRow()
    if s != 1:
        if s == ws.max_row:
            txt2.delete(0, END)
            txt2.insert(0, str(ws.cell(row=2, column=2).value))
            lb_student.config(text="1/"+str(ws.max_row-1)+" 번째 학생")
        else:
            txt2.delete(0, END)
            txt2.insert(0, str(ws.cell(row=s+1, column=2).value))
            lb_student.config(text=str(s)+"/"+str(ws.max_row-1)+" 번째 학생")
    find_score()

def stum():
    s = StuRow()
    if s != 1:
        if s == 2:
            txt2.delete(0, END)
            txt2.insert(0, str(ws.cell(row=ws.max_row, column=2).value))
            lb_student.config(text=str(ws.max_row-1)+"/"+str(ws.max_row-1)+" 번째 학생")
        else:
            txt2.delete(0, END)
            txt2.insert(0, str(ws.cell(row=s-1, column=2).value))
            lb_student.config(text=str(s-2)+"/"+str(ws.max_row-1)+" 번째 학생")
    find_score()

def studel():
    response = msgbox.askokcancel("경고", "학생의 정보가 모두 삭제됩니다. \n계속 하시겠습니까?")
    if response == 1:
        ws.delete_rows(StuRow())
    elif response == 0:
        pass

#레포트 번호 바꾸기 버튼
btn_weekp = Button(root, text=">>", command=weekp)
btn_weekm = Button(root, text="<<", command=weekm)

#레포트 번호 표시
lb_week = Label(root, text="1/9 주차 레포트")

#학생 번호 표시
if ws.max_row == 1:
    lb_student = Label(root, text="새 파일입니다")
else:
    lb_student = Label(root, text="1/"+str(ws.max_row-1)+" 번째 학생")

#학생 바꾸기 버튼
btn_stup = Button(root, text=">>", command=stup)
btn_stum = Button(root, text="<<", command=stum)
#학생 삭제 버튼
btn_delstu = Button(root, text="학생삭제",command=studel)
#학생 조회 버튼
btn_student = Button(root, text="학번조회", command=find_score)
#입력버튼
btn_save = Button(root, text="점수저장", command=save_score)


lb_week.grid(row=0, column=1)        #  레포트 번호 레이블

btn_weekp.grid(row=0, column=2)      #  레포트 번호 +1 버튼
btn_weekm.grid(row=0, column=0)      #  레포트 번호 -1 버튼

lb_student.grid(row=1, column=1)     #  학생 번호

btn_stup.grid(row=1, column=2)       #  학생 번호 +1 버튼
btn_stum.grid(row=1, column=0)       #  학생 번호 -1 버튼

lb1.grid(row=2, column=0)            #  이름칸 레이블
txt1.grid(row=2, column=1)           #  이름칸 입력칸

lb2.grid(row=3, column=0)            #  학번칸 레이블
txt2.grid(row=3, column=1)           #  학번칸 입력칸

lb3.grid(row=4, column=0)           #  점수칸 레이블
txt3.grid(row=4, column=1)           #  점수칸 입력칸

btn_delstu.grid(row=2, column=2)
btn_student.grid(row=3, column=2)    #  학번 조회 버튼
btn_save.grid(row=4, column=2)       #  점수 저장 버튼

wb.save("실험수업점수표.xlsx")
wb.close()

root.mainloop()
